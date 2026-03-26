"""
AMC Portfolio Disclosure Parser
Parses monthly portfolio Excel files (e.g., Axis AMC) and upserts
equity holdings into fund_top_holdings table.

Usage:
  python3 scripts/import-amc-holdings.py <path-to-excel>

Excel structure (per AMC disclosure standard):
  - "Index" sheet: Sr No | Short Name | Scheme Name
  - Per-fund sheets: Row 1 = code + name, Row 3 = statement date,
    Row 4 = headers, Rows 5+ = holdings grouped by asset type
"""

import sys
import os
import re
from datetime import datetime
import openpyxl
import psycopg2
from psycopg2.extras import execute_values

DB_URL = os.environ.get(
    "DATABASE_URL",
    "postgresql://postgres:zZGzhpULOgKqXvnnutWEjCengioSheMD@turntable.proxy.rlwy.net:19665/railway"
)


def parse_statement_date(ws):
    """Extract the 'as on' date from row 3."""
    for row in ws.iter_rows(min_row=2, max_row=5, values_only=True):
        for cell in row:
            if cell and isinstance(cell, str) and "as on" in cell.lower():
                # "Monthly Portfolio Statement as on January 31, 2026"
                m = re.search(r"as on\s+(.+)", cell, re.IGNORECASE)
                if m:
                    date_str = m.group(1).strip()
                    for fmt in ("%B %d, %Y", "%d %B %Y", "%d-%b-%Y", "%b %d, %Y"):
                        try:
                            return datetime.strptime(date_str, fmt).date()
                        except ValueError:
                            continue
    return None


def parse_equity_holdings(ws):
    """
    Extract equity holdings from a fund sheet.
    Returns list of dicts: {name, isin, sector, weight_pct, quantity, market_value_lakhs}
    """
    holdings = []
    in_equity_section = False
    in_listed = False

    for row in ws.iter_rows(min_row=5, values_only=False):
        cells = {c.column: c.value for c in row}
        b_val = cells.get(2, "")  # Column B
        if b_val is None:
            b_val = ""

        b_str = str(b_val).strip()

        # Detect section boundaries
        if "equity & equity related" in b_str.lower():
            in_equity_section = True
            continue
        if in_equity_section and "(a)" in b_str.lower() and "listed" in b_str.lower():
            in_listed = True
            continue
        if in_equity_section and b_str.lower().startswith("sub total"):
            if in_listed:
                in_listed = False
                continue
        if b_str.lower().startswith("total") and in_equity_section and not in_listed:
            in_equity_section = False
            continue

        # Skip non-equity sections
        if not in_equity_section or not in_listed:
            continue

        # Skip sub-headers and empty rows
        if "(b)" in b_str.lower() or "unlisted" in b_str.lower():
            in_listed = False
            continue

        name = cells.get(2)  # B = Name
        isin = cells.get(3)  # C = ISIN
        sector = cells.get(4)  # D = Industry/Rating
        quantity = cells.get(5)  # E = Quantity
        market_value = cells.get(6)  # F = Market Value (Lakhs)
        weight = cells.get(7)  # G = % to Net Assets

        # Must have name + weight to be a valid holding
        if not name or weight is None:
            continue
        if isinstance(name, str) and name.strip().lower() in ("sub total", "total", "nil", ""):
            continue

        try:
            weight_pct = float(weight) * 100  # Convert 0.0879 → 8.79
        except (ValueError, TypeError):
            continue

        if weight_pct <= 0:
            continue

        holdings.append({
            "name": str(name).strip().rstrip(" **").rstrip("**"),
            "isin": str(isin).strip() if isin else None,
            "sector": str(sector).strip() if sector else None,
            "weight_pct": round(weight_pct, 2),
            "quantity": int(quantity) if quantity else None,
            "market_value_lakhs": round(float(market_value), 2) if market_value else None,
        })

    # Sort by weight descending, take top 10
    holdings.sort(key=lambda h: h["weight_pct"], reverse=True)
    return holdings[:10]


def match_fund_to_scheme_codes(cur, fund_name):
    """
    Match an AMC fund name (e.g. 'Axis Large Cap Fund') to scheme_codes in DB.
    Returns list of scheme_codes (Direct Growth preferred, then all variants).
    """
    # Clean up the fund name for matching
    clean = fund_name.strip()
    # Try exact substring match
    cur.execute(
        "SELECT scheme_code FROM funds WHERE scheme_name ILIKE %s AND is_active = true ORDER BY scheme_code",
        [f"%{clean}%"]
    )
    codes = [r[0] for r in cur.fetchall()]
    if codes:
        return codes

    # Try with key words only (remove parenthetical parts)
    clean = re.sub(r"\s*\(.*?\)", "", clean).strip()
    cur.execute(
        "SELECT scheme_code FROM funds WHERE scheme_name ILIKE %s AND is_active = true ORDER BY scheme_code",
        [f"%{clean}%"]
    )
    return [r[0] for r in cur.fetchall()]


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 import-amc-holdings.py <path-to-excel>")
        sys.exit(1)

    excel_path = sys.argv[1]
    if not os.path.exists(excel_path):
        print(f"File not found: {excel_path}")
        sys.exit(1)

    print(f"Loading {excel_path}...")
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # Build index: short_name -> fund_name
    index_ws = wb["Index"]
    fund_index = {}
    for row in index_ws.iter_rows(min_row=2, values_only=True):
        sr, short_name, scheme_name = row[0], row[1], row[2]
        if short_name and scheme_name:
            fund_index[str(short_name).strip()] = str(scheme_name).strip()

    print(f"Found {len(fund_index)} funds in Index sheet")

    conn = psycopg2.connect(DB_URL, connect_timeout=15)
    cur = conn.cursor()

    total_imported = 0
    total_skipped = 0
    funds_matched = 0

    for sheet_name in wb.sheetnames:
        if sheet_name == "Index":
            continue

        ws = wb[sheet_name]

        # Get fund name from row 1 or index
        fund_name = fund_index.get(sheet_name)
        if not fund_name:
            cell_b1 = ws["B1"].value
            if cell_b1:
                fund_name = str(cell_b1).strip()
            else:
                continue

        # Parse statement date
        as_of_date = parse_statement_date(ws)
        if not as_of_date:
            print(f"  [{sheet_name}] No statement date found, skipping")
            total_skipped += 1
            continue

        # Parse equity holdings
        holdings = parse_equity_holdings(ws)
        if not holdings:
            total_skipped += 1
            continue

        # Match to scheme_codes in DB
        scheme_codes = match_fund_to_scheme_codes(cur, fund_name)
        if not scheme_codes:
            print(f"  [{sheet_name}] '{fund_name}' - no DB match, skipping")
            total_skipped += 1
            continue

        funds_matched += 1
        print(f"  [{sheet_name}] '{fund_name}' -> {len(scheme_codes)} variants, {len(holdings)} holdings")

        # Upsert holdings for all scheme_code variants
        for sc in scheme_codes:
            # Delete existing holdings for this scheme_code + date
            cur.execute(
                "DELETE FROM fund_top_holdings WHERE scheme_code = %s AND as_of_date = %s",
                [sc, as_of_date]
            )
            # Insert new holdings
            rows = []
            for rank, h in enumerate(holdings, 1):
                rows.append((
                    sc,
                    h["name"],
                    h["isin"],
                    h["weight_pct"],
                    h["sector"],
                    rank,
                    as_of_date,
                ))

            execute_values(
                cur,
                """INSERT INTO fund_top_holdings
                   (scheme_code, holding_name, holding_isin, weight_pct, sector, rank, as_of_date)
                   VALUES %s""",
                rows
            )
            total_imported += len(rows)

    conn.commit()
    cur.close()
    conn.close()

    print(f"\nDone! {funds_matched} funds matched, {total_imported} holdings imported, {total_skipped} sheets skipped")


if __name__ == "__main__":
    main()
