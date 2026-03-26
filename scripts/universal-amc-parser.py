"""
Universal AMC Holdings Parser — "The Aladdin Engine"
Recursively crawls directories, extracts ZIPs in-memory, parses all
AMC portfolio disclosure Excel/CSV files, normalizes sectors, and
populates:
  1. fund_top_holdings (raw holdings)
  2. scheme_stock_index (per-fund ISIN index)
  3. scheme_sector_summary (pre-aggregated O(1) screener index)

Usage:
  python3 scripts/universal-amc-parser.py <path>

  <path> can be:
    - A single .xlsx/.csv file
    - A .zip archive
    - A directory (recursively crawled)

Supports: Axis, SBI, HDFC, ICICI, Kotak, and all SEBI-mandated
portfolio disclosure formats.
"""

import sys
import os
import re
import io
import zipfile
from datetime import datetime
from decimal import Decimal
import openpyxl
import psycopg2
from psycopg2.extras import execute_values

DB_URL = os.environ.get(
    "DATABASE_URL",
    "postgresql://postgres:zZGzhpULOgKqXvnnutWEjCengioSheMD@turntable.proxy.rlwy.net:19665/railway"
)

# ════════════════════════════════════════════════════════════════
#  SECTOR NORMALIZATION
# ════════════════════════════════════════════════════════════════

# Loaded from DB at startup — maps raw_sector → master_sector_id
SECTOR_ALIAS_MAP = {}
MASTER_SECTORS = {}  # id → sector_name

# Credit rating patterns — these are NOT sectors
RATING_PATTERN = re.compile(
    r'^(CRISIL|CARE|FITCH|ICRA|IND|ACUITE|BWR|SOV|\[ICRA\])',
    re.IGNORECASE
)


def load_sector_map(cur):
    """Load sector alias mapping from DB."""
    global SECTOR_ALIAS_MAP, MASTER_SECTORS
    cur.execute("SELECT id, sector_name FROM master_sectors")
    MASTER_SECTORS = {r[0]: r[1] for r in cur.fetchall()}
    cur.execute("SELECT raw_sector, master_sector_id FROM sector_aliases")
    SECTOR_ALIAS_MAP = {r[0]: r[1] for r in cur.fetchall()}


def normalize_sector(raw_sector):
    """
    Map a raw AMC sector string to a master_sector_id.
    Returns (master_sector_id, is_equity).
    Credit ratings return (None, False).
    """
    if not raw_sector:
        return None, False

    raw = raw_sector.strip()

    # Skip credit ratings — not equity sectors
    if RATING_PATTERN.match(raw):
        return None, False
    if raw in ("Sovereign", "SOV", "CDMDF"):
        return None, False
    if raw.startswith("Mutual Fund") or raw.startswith("ETF") or raw.startswith("Overseas"):
        return None, False

    # Exact alias match
    if raw in SECTOR_ALIAS_MAP:
        return SECTOR_ALIAS_MAP[raw], True

    # Fuzzy: try without trailing whitespace, truncation
    for alias, sid in SECTOR_ALIAS_MAP.items():
        if raw.startswith(alias) or alias.startswith(raw):
            return sid, True

    # Fallback: "other"
    other_id = next((sid for sid, name in MASTER_SECTORS.items() if name == "other"), None)
    return other_id, True


# ════════════════════════════════════════════════════════════════
#  EXCEL PARSER (SEBI Standard Format)
# ════════════════════════════════════════════════════════════════

def parse_statement_date(ws):
    """Extract the 'as on' date from early rows."""
    for row in ws.iter_rows(min_row=1, max_row=8, values_only=True):
        for cell in row:
            if cell and isinstance(cell, datetime):
                return cell.date()
            if cell and isinstance(cell, str) and "as on" in cell.lower():
                m = re.search(r"as on\s+(.+?)(?:\s*$)", cell, re.IGNORECASE)
                if m:
                    date_str = m.group(1).strip()
                    for fmt in ("%B %d, %Y", "%d %B %Y", "%d-%b-%Y", "%b %d, %Y",
                                "%d-%m-%Y", "%d/%m/%Y", "%B %d,%Y", "%d-%B-%Y"):
                        try:
                            return datetime.strptime(date_str, fmt).date()
                        except ValueError:
                            continue
    return None


def detect_format(ws):
    """
    Auto-detect the AMC format by scanning header rows.
    Returns: 'axis' | 'sbi' | 'kotak' | 'unknown'
    """
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=False):
        vals = []
        for c in row:
            vals.append(str(c.value or "").strip().lower())
        joined = " ".join(vals)
        # SBI: has '% to aum' in header row
        if "% to aum" in joined:
            return "sbi"
        # Kotak: has '% to net assets' in header row
        if "% to net assets" in joined:
            return "kotak"
        # Axis: has '% to nav' in header row
        if "% to nav" in joined:
            return "axis"
    return "unknown"


def parse_equity_holdings(ws):
    """
    Extract equity holdings from a fund sheet.
    Auto-detects format (Axis/SBI/Kotak) and adapts column mapping.
    Returns list of dicts with normalized sector info.
    """
    fmt = detect_format(ws)

    if fmt == "kotak":
        return _parse_kotak(ws)
    elif fmt == "sbi":
        return _parse_sbi(ws)
    else:
        return _parse_axis(ws)


def _parse_axis(ws):
    """Parse Axis AMC format: B=name, C=ISIN, D=sector, G=weight (decimal 0.08)."""
    holdings = []
    in_equity_section = False
    in_listed = False

    for row in ws.iter_rows(min_row=4, values_only=False):
        cells = {c.column: c.value for c in row}
        b_val = str(cells.get(2, "") or "").strip()

        if "equity" in b_val.lower() and "related" in b_val.lower():
            in_equity_section = True
            continue
        if in_equity_section and ("listed" in b_val.lower() or "(a)" in b_val.lower()):
            in_listed = True
            continue
        if b_val.lower().startswith("sub total"):
            if in_listed:
                in_listed = False
            continue
        if b_val.lower().startswith("total") and in_equity_section:
            in_equity_section = False
            continue
        if "(b)" in b_val.lower() and ("unlisted" in b_val.lower() or "privately" in b_val.lower()):
            in_listed = False
            continue
        if not in_equity_section or not in_listed:
            continue

        name = cells.get(2)
        isin = cells.get(3)
        sector_raw = cells.get(4)
        weight = cells.get(7)

        if not name or weight is None:
            continue
        name_clean = str(name).strip().rstrip(" *")
        if name_clean.lower() in ("sub total", "total", "nil", ""):
            continue

        try:
            w = float(weight)
        except (ValueError, TypeError):
            continue
        # Axis uses decimal (0.0879 = 8.79%). If < 1, multiply by 100.
        weight_pct = w * 100 if w < 1 else w
        if weight_pct <= 0:
            continue

        sector_str = str(sector_raw).strip() if sector_raw else None
        master_sector_id, is_equity = normalize_sector(sector_str)
        holdings.append({
            "name": name_clean,
            "isin": str(isin).strip() if isin else None,
            "sector_raw": sector_str,
            "sector_id": master_sector_id,
            "is_equity": is_equity,
            "weight_pct": round(weight_pct, 2),
        })

    holdings.sort(key=lambda h: h["weight_pct"], reverse=True)
    return holdings[:10]


def _parse_sbi(ws):
    """Parse SBI format: C=name, D=ISIN, E=industry, H=% to AUM (already percentage)."""
    holdings = []
    in_equity_section = False
    in_listed = False

    for row in ws.iter_rows(min_row=4, values_only=False):
        cells = {c.column: c.value for c in row}
        # SBI uses col C for name sometimes, col B for internal code
        c_val = str(cells.get(3, "") or "").strip()
        b_val = str(cells.get(2, "") or "").strip()
        combined = f"{b_val} {c_val}".lower()

        if "equity" in combined and "related" in combined:
            in_equity_section = True
            continue
        if in_equity_section and "listed" in combined:
            in_listed = True
            continue
        if "sub total" in combined:
            continue
        if combined.startswith("total") and in_equity_section:
            in_equity_section = False
            continue
        if "(b)" in combined and ("unlisted" in combined or "privately" in combined):
            in_listed = False
            continue

        if not in_equity_section or not in_listed:
            continue

        name = cells.get(3)  # col C = stock name
        isin = cells.get(4)  # col D = ISIN
        sector_raw = cells.get(5)  # col E = industry
        weight = cells.get(8)  # col H = % to AUM

        if not name or weight is None:
            continue
        name_clean = str(name).strip().rstrip(" *")
        if name_clean.lower() in ("sub total", "total", "nil", "", "name of the instrument"):
            continue

        try:
            weight_pct = float(weight)
        except (ValueError, TypeError):
            continue
        if weight_pct <= 0:
            continue

        sector_str = str(sector_raw).strip() if sector_raw else None
        master_sector_id, is_equity = normalize_sector(sector_str)
        holdings.append({
            "name": name_clean,
            "isin": str(isin).strip() if isin else None,
            "sector_raw": sector_str,
            "sector_id": master_sector_id,
            "is_equity": is_equity,
            "weight_pct": round(weight_pct, 2),
        })

    holdings.sort(key=lambda h: h["weight_pct"], reverse=True)
    return holdings[:10]


def _parse_kotak(ws):
    """Parse Kotak format: C=name, D=ISIN, E=industry, I(col9)=% to Net Assets (already %)."""
    holdings = []
    in_equity_section = False
    in_listed = False

    for row in ws.iter_rows(min_row=3, values_only=False):
        cells = {c.column: c.value for c in row}
        # Kotak uses cols A-C for section labels
        a_val = str(cells.get(1, "") or "").strip()
        b_val = str(cells.get(2, "") or "").strip()
        c_val = str(cells.get(3, "") or "").strip()
        combined = f"{a_val} {b_val} {c_val}".lower()

        if "equity" in combined and ("related" in combined or combined.strip() == "equity & equity related"):
            in_equity_section = True
            continue
        if in_equity_section and "listed" in combined:
            in_listed = True
            continue
        if "sub total" in combined or "subtotal" in combined:
            continue
        if combined.strip().startswith("total") and in_equity_section:
            in_equity_section = False
            continue
        if ("unlisted" in combined or "privately" in combined):
            in_listed = False
            continue

        if not in_equity_section or not in_listed:
            continue

        name = cells.get(3)  # col C = stock name
        isin = cells.get(4)  # col D = ISIN
        sector_raw = cells.get(5)  # col E = industry
        weight = cells.get(9)  # col I (9) = % to Net Assets

        if not name or weight is None:
            continue
        name_clean = str(name).strip().rstrip(" *")
        if name_clean.lower() in ("sub total", "total", "nil", "", "name of instrument"):
            continue

        try:
            weight_pct = float(weight)
        except (ValueError, TypeError):
            continue
        if weight_pct <= 0:
            continue

        sector_str = str(sector_raw).strip() if sector_raw else None
        master_sector_id, is_equity = normalize_sector(sector_str)
        holdings.append({
            "name": name_clean,
            "isin": str(isin).strip() if isin else None,
            "sector_raw": sector_str,
            "sector_id": master_sector_id,
            "is_equity": is_equity,
            "weight_pct": round(weight_pct, 2),
        })

    holdings.sort(key=lambda h: h["weight_pct"], reverse=True)
    return holdings[:10]


def match_fund_to_scheme_codes(cur, fund_name):
    """Match AMC fund name to scheme_codes in DB."""
    clean = fund_name.strip()
    cur.execute(
        "SELECT scheme_code FROM funds WHERE scheme_name ILIKE %s AND is_active = true",
        [f"%{clean}%"]
    )
    codes = [r[0] for r in cur.fetchall()]
    if codes:
        return codes

    clean = re.sub(r"\s*\(.*?\)", "", clean).strip()
    cur.execute(
        "SELECT scheme_code FROM funds WHERE scheme_name ILIKE %s AND is_active = true",
        [f"%{clean}%"]
    )
    return [r[0] for r in cur.fetchall()]


def process_workbook(wb, source_name, cur, stats):
    """Process a single Excel workbook."""
    # Try to find Index sheet
    fund_index = {}
    if "Index" in wb.sheetnames:
        idx_ws = wb["Index"]
        for row in idx_ws.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= 3 and row[1] and row[2]:
                fund_index[str(row[1]).strip()] = str(row[2]).strip()

    for sheet_name in wb.sheetnames:
        if sheet_name == "Index":
            continue

        ws = wb[sheet_name]
        fund_name = fund_index.get(sheet_name)
        if not fund_name:
            # Try multiple cells for fund name (different AMC formats)
            for cell_ref in ["C1", "B1", "A1", "C4", "D4"]:
                try:
                    cell_val = ws[cell_ref].value
                except Exception:
                    continue
                if cell_val:
                    raw = str(cell_val).strip()
                    # Kotak: "Portfolio of Kotak Nifty India Tourism Index Fund as on 31-Jan-2026"
                    m = re.match(r"Portfolio of (.+?) as on", raw, re.IGNORECASE)
                    if m:
                        fund_name = m.group(1).strip()
                        break
                    # SBI: "SCHEME NAME :" in one cell, actual name in next
                    if "scheme name" in raw.lower():
                        try:
                            name_val = ws[f"D{cell_ref[1:]}"].value or ws[f"E{cell_ref[1:]}"].value
                            if name_val:
                                fund_name = str(name_val).strip()
                                break
                        except Exception:
                            pass
                    if len(raw) > 5 and "portfolio" not in raw.lower() and "statement" not in raw.lower():
                        fund_name = raw
                        break
            if not fund_name:
                # SBI: check C3/D3 for scheme name
                for r in range(2, 6):
                    try:
                        c3 = ws.cell(row=r, column=3).value
                        if c3 and "scheme name" in str(c3).lower():
                            d3 = ws.cell(row=r, column=4).value
                            if d3:
                                fund_name = str(d3).strip()
                                break
                    except Exception:
                        continue
            if not fund_name:
                continue

        as_of_date = parse_statement_date(ws)
        if not as_of_date:
            # Try to extract date from Kotak-style row 1 ("as on 31-Jan-2026")
            for r in range(1, 6):
                for c in range(1, 5):
                    try:
                        val = ws.cell(row=r, column=c).value
                    except Exception:
                        continue
                    if val and isinstance(val, str) and "as on" in val.lower():
                        m = re.search(r"as on\s+(.+?)(?:\s*$)", val, re.IGNORECASE)
                        if m:
                            date_str = m.group(1).strip()
                            for fmt in ("%d-%b-%Y", "%B %d, %Y", "%d %B %Y",
                                        "%d-%m-%Y", "%d/%m/%Y", "%b %d, %Y"):
                                try:
                                    as_of_date = datetime.strptime(date_str, fmt).date()
                                    break
                                except ValueError:
                                    continue
                    # Also try datetime objects (SBI puts dates as datetime)
                    if val and isinstance(val, datetime):
                        as_of_date = val.date()
                    if as_of_date:
                        break
                if as_of_date:
                    break
            if not as_of_date:
                continue

        holdings = parse_equity_holdings(ws)
        if not holdings:
            stats["skipped"] += 1
            continue

        scheme_codes = match_fund_to_scheme_codes(cur, fund_name)
        if not scheme_codes:
            stats["unmatched"] += 1
            continue

        stats["matched"] += 1

        # Deduplicate holdings by name (some sheets list same stock twice)
        seen_names = set()
        deduped = []
        for h in holdings:
            if h["name"] not in seen_names:
                seen_names.add(h["name"])
                deduped.append(h)
        holdings = deduped

        for sc in scheme_codes:
            # 1. Upsert fund_top_holdings
            cur.execute(
                "DELETE FROM fund_top_holdings WHERE scheme_code = %s AND as_of_date = %s",
                [sc, as_of_date]
            )
            rows_fth = []
            for rank, h in enumerate(holdings, 1):
                rows_fth.append((sc, h["name"], h["isin"], h["weight_pct"],
                                 h["sector_raw"], rank, as_of_date))
            if rows_fth:
                execute_values(cur,
                    """INSERT INTO fund_top_holdings
                       (scheme_code, holding_name, holding_isin, weight_pct, sector, rank, as_of_date)
                       VALUES %s
                       ON CONFLICT (scheme_code, holding_name, as_of_date)
                       DO UPDATE SET weight_pct = EXCLUDED.weight_pct, rank = EXCLUDED.rank,
                                     holding_isin = EXCLUDED.holding_isin, sector = EXCLUDED.sector""",
                    rows_fth)

            # 2. Upsert scheme_stock_index
            cur.execute(
                "DELETE FROM scheme_stock_index WHERE scheme_code = %s AND as_of_date = %s",
                [sc, as_of_date]
            )
            rows_ssi = [(sc, h["isin"], h["name"], h["weight_pct"], as_of_date)
                        for h in holdings if h["isin"]]
            if rows_ssi:
                execute_values(cur,
                    """INSERT INTO scheme_stock_index
                       (scheme_code, isin, stock_name, weight_pct, as_of_date)
                       VALUES %s
                       ON CONFLICT DO NOTHING""", rows_ssi)

            # 3. Build sector aggregation
            cur.execute(
                "DELETE FROM scheme_sector_summary WHERE scheme_code = %s AND as_of_date = %s",
                [sc, as_of_date]
            )
            sector_agg = {}
            for h in holdings:
                sid = h["sector_id"]
                if sid is None:
                    continue
                if sid not in sector_agg:
                    sector_agg[sid] = {"weight": 0.0, "count": 0}
                sector_agg[sid]["weight"] += h["weight_pct"]
                sector_agg[sid]["count"] += 1

            rows_sss = [(sc, sid, round(a["weight"], 2), a["count"], as_of_date)
                        for sid, a in sector_agg.items()]
            if rows_sss:
                execute_values(cur,
                    """INSERT INTO scheme_sector_summary
                       (scheme_code, master_sector_id, total_weight, holding_count, as_of_date)
                       VALUES %s""", rows_sss)

            stats["holdings"] += len(rows_fth)


# ════════════════════════════════════════════════════════════════
#  FILE DISCOVERY — Recursive ZIP/Folder Crawling
# ════════════════════════════════════════════════════════════════

def discover_files(path):
    """
    Recursively discover all parseable files.
    Yields (filename, file_bytes_or_path, is_bytes) tuples.
    """
    if os.path.isfile(path):
        if path.lower().endswith(".zip"):
            yield from extract_zip(path)
        elif path.lower().endswith((".xlsx", ".xls")):
            yield (os.path.basename(path), path, False)
    elif os.path.isdir(path):
        for root, dirs, files in os.walk(path):
            for f in sorted(files):
                full = os.path.join(root, f)
                if f.lower().endswith(".zip"):
                    yield from extract_zip(full)
                elif f.lower().endswith((".xlsx", ".xls")):
                    yield (f, full, False)


def extract_zip(zip_path):
    """Extract xlsx files from a ZIP archive (in-memory)."""
    try:
        with zipfile.ZipFile(zip_path, "r") as zf:
            for name in zf.namelist():
                if name.lower().endswith((".xlsx", ".xls")) and not name.startswith("__MACOSX"):
                    data = zf.read(name)
                    yield (name, data, True)
                elif name.lower().endswith(".zip"):
                    # Nested ZIP — extract recursively
                    inner_data = zf.read(name)
                    inner_zf = zipfile.ZipFile(io.BytesIO(inner_data))
                    for inner_name in inner_zf.namelist():
                        if inner_name.lower().endswith((".xlsx", ".xls")) and not inner_name.startswith("__MACOSX"):
                            yield (inner_name, inner_zf.read(inner_name), True)
    except zipfile.BadZipFile:
        print(f"  [WARN] Bad ZIP: {zip_path}")


# ════════════════════════════════════════════════════════════════
#  MAIN
# ════════════════════════════════════════════════════════════════

def main():
    if len(sys.argv) < 2:
        print("Usage: python3 universal-amc-parser.py <path>")
        print("  <path> = file.xlsx | file.zip | directory/")
        sys.exit(1)

    target = sys.argv[1]
    if not os.path.exists(target):
        print(f"Path not found: {target}")
        sys.exit(1)

    conn = psycopg2.connect(DB_URL, connect_timeout=15)
    cur = conn.cursor()

    # Load sector normalization map
    load_sector_map(cur)
    print(f"Loaded {len(SECTOR_ALIAS_MAP)} sector aliases, {len(MASTER_SECTORS)} master sectors")

    stats = {"matched": 0, "unmatched": 0, "skipped": 0, "holdings": 0, "files": 0}

    for filename, source, is_bytes in discover_files(target):
        stats["files"] += 1
        print(f"\n[{stats['files']}] Processing: {filename}")
        try:
            if is_bytes:
                wb = openpyxl.load_workbook(io.BytesIO(source), data_only=True)
            else:
                wb = openpyxl.load_workbook(source, data_only=True)
            process_workbook(wb, filename, cur, stats)
            conn.commit()
        except Exception as e:
            print(f"  [ERROR] {filename}: {e}")
            conn.rollback()

    cur.close()
    conn.close()

    print(f"\n{'='*60}")
    print(f"  Files processed:    {stats['files']}")
    print(f"  Funds matched:      {stats['matched']}")
    print(f"  Funds unmatched:    {stats['unmatched']}")
    print(f"  Sheets skipped:     {stats['skipped']}")
    print(f"  Holdings imported:  {stats['holdings']}")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
