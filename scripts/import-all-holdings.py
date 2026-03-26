"""
Comprehensive AMC Portfolio Holdings Importer
Handles: ICICI (per-fund files), SBI, Quant, Kotak, Axis-style
Upserts into fund_top_holdings table.

Usage:
  python3 scripts/import-all-holdings.py [--dry-run]
"""

from __future__ import annotations
import os
import sys
import re
import psycopg2
from psycopg2.extras import execute_values
from datetime import datetime, date as DateType
from pathlib import Path
from typing import Optional
import openpyxl

# ── Config ───────────────────────────────────────────────────
DB_URL = os.environ.get(
    "DATABASE_URL",
    "postgresql://postgres:zZGzhpULOgKqXvnnutWEjCengioSheMD@turntable.proxy.rlwy.net:19665/railway"
)
BASE_DIR = Path(__file__).parent.parent / "Portfolio_AMC_Monthly"
DRY_RUN = "--dry-run" in sys.argv

# ── Helpers ──────────────────────────────────────────────────

def parse_date(raw: str | None) -> Optional[DateType]:
    if not raw:
        return None
    if isinstance(raw, datetime):
        return DateType(raw.year, raw.month, raw.day)
    raw = str(raw).strip()
    patterns = [
        r"(\d{1,2})[- /](\d{1,2})[- /](\d{4})",         # dd/mm/yyyy
        r"(\d{4})-(\d{2})-(\d{2})",                       # yyyy-mm-dd
    ]
    # Named month patterns
    month_pats = [
        r"(?:as on|as at|dated?)[:\s]+(.+)",              # "as on Feb 28, 2026"
        r"(?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{1,2},\s+\d{4}",
    ]
    for pat in month_pats:
        m = re.search(pat, raw, re.IGNORECASE)
        if m:
            date_str = m.group(1).strip() if m.lastindex else m.group(0).strip()
            for fmt in ("%B %d, %Y", "%d %B %Y", "%B %d,%Y", "%b %d, %Y", "%b %d,%Y"):
                try:
                    return datetime.strptime(date_str, fmt).date()  # type: ignore
                except ValueError:
                    continue
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(raw[:10], fmt).date()  # type: ignore
        except ValueError:
            continue
    return None


def is_valid_isin(val: Optional[str]) -> bool:
    if not val:
        return False
    s = str(val).strip()
    return bool(re.match(r'^IN[A-Z0-9]{10}$', s))


def to_float(val) -> Optional[float]:
    try:
        return float(str(val).replace(',', '').replace('%', '').strip())
    except (ValueError, TypeError, AttributeError):
        return None


def clean_name(val) -> str:
    if not val:
        return ''
    return str(val).strip()


def match_scheme_codes(cur, fund_name: str) -> list[str]:
    """Fuzzy-match a fund name to scheme_codes. Returns all matching variants."""
    if not fund_name:
        return []
    # Strip plan/option suffixes for matching
    clean = re.sub(r'\s*[-–]\s*(Direct|Regular)\s*(Plan)?\s*', '', fund_name, flags=re.IGNORECASE)
    clean = re.sub(r'\s*[-–]\s*(Growth|IDCW|Dividend)(\s+Option|\s+Plan|\s+Payout|\s+Reinvestment)?\s*$', '', clean, flags=re.IGNORECASE)
    clean = re.sub(r'\s*(Plan|Option|Fund)\s*$', '', clean, flags=re.IGNORECASE).strip()

    # Exact match first
    cur.execute("SELECT scheme_code FROM funds WHERE scheme_name ILIKE %s AND is_active IS NOT FALSE", [fund_name])
    exact = [r[0] for r in cur.fetchall()]
    if exact:
        return exact

    # Starts-with match on cleaned name
    if len(clean) > 8:
        cur.execute(
            "SELECT scheme_code FROM funds WHERE scheme_name ILIKE %s AND is_active IS NOT FALSE",
            [f"{clean}%"]
        )
        rows = [r[0] for r in cur.fetchall()]
        if rows:
            return rows

    # Trigram similarity
    cur.execute(
        """SELECT scheme_code FROM funds
           WHERE similarity(scheme_name, %s) > 0.4
             AND is_active IS NOT FALSE
           ORDER BY similarity(scheme_name, %s) DESC LIMIT 10""",
        [clean, clean]
    )
    return [r[0] for r in cur.fetchall()]


def upsert_holdings(cur, scheme_code: str, holdings: list[dict], as_of_date):
    if not holdings or not as_of_date:
        return 0
    # Deduplicate by name — aggregate weight, keep first ISIN
    seen: dict[str, dict] = {}
    for h in holdings:
        name = h['name']
        if name not in seen:
            seen[name] = dict(h)
        else:
            seen[name]['weight'] = (seen[name]['weight'] or 0) + (h['weight'] or 0)
    deduped = sorted(seen.values(), key=lambda x: x['weight'] or 0, reverse=True)[:20]

    cur.execute(
        "DELETE FROM fund_top_holdings WHERE scheme_code = %s AND as_of_date = %s",
        [scheme_code, as_of_date]
    )
    rows = [
        (scheme_code, h['name'], h.get('isin'), h['weight'], h.get('sector'), i + 1, as_of_date)
        for i, h in enumerate(deduped)
    ]
    execute_values(cur,
        """INSERT INTO fund_top_holdings
           (scheme_code, holding_name, holding_isin, weight_pct, sector, rank, as_of_date)
           VALUES %s""",
        rows
    )
    return len(rows)


# ─────────────────────────────────────────────────────────────
# ICICI — per-fund .xlsx files in a directory
# ─────────────────────────────────────────────────────────────
def parse_icici_file(filepath: Path) -> tuple[Optional[str], Optional[DateType], list[dict]]:
    """
    Row 2 col[1] = fund name
    Row 3 col[1] = "Portfolio as on Feb 28,2026"
    Row 4        = headers
    Row 5+       = holdings (name=col[1], isin=col[2], sector=col[4], weight=col[7])
    """
    try:
        wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(max_row=200, values_only=True))
        wb.close()
    except Exception as e:
        print(f"  ICICI read error {filepath.name}: {e}")
        return None, None, []

    fund_name = clean_name(rows[1][1]) if len(rows) > 1 and rows[1][1] else None
    as_of_date = None
    if len(rows) > 2 and rows[2][1]:
        as_of_date = parse_date(str(rows[2][1]))

    holdings = []
    for row in rows[4:]:
        if len(row) < 8:
            continue
        name  = clean_name(row[1])
        isin  = clean_name(row[2])
        sector= clean_name(row[4]) if len(row) > 4 else ''
        weight= to_float(row[7]) if len(row) > 7 else None

        if not is_valid_isin(isin) or weight is None:
            continue
        # ICICI stores weights as decimal fractions (0.094 = 9.4%)
        if weight > 0 and weight < 1:
            weight = round(weight * 100, 4)
        if weight <= 0 or weight > 100:
            continue
        holdings.append({'name': name, 'isin': isin, 'sector': sector, 'weight': weight})
    return fund_name, as_of_date, holdings


def import_icici(cur) -> tuple[int, int]:  # type: ignore
    icici_dir = BASE_DIR / "Portfolio_Monthly_ICICI"
    if not icici_dir.exists():
        print("ICICI directory not found, skipping")
        return 0, 0
    imported, skipped = 0, 0
    for xlsx_file in sorted(icici_dir.glob("*.xlsx")):
        fund_name, as_of_date, holdings = parse_icici_file(xlsx_file)
        if not holdings:
            skipped += 1
            continue
        codes = match_scheme_codes(cur, fund_name or xlsx_file.stem)
        if not codes:
            print(f"  ICICI '{fund_name}' — no DB match")
            skipped += 1
            continue
        print(f"  ICICI '{fund_name}' → {len(codes)} variants, {len(holdings)} holdings, date={as_of_date}")
        for sc in codes:
            imported += upsert_holdings(cur, sc, holdings, as_of_date)
    return imported, skipped


# ─────────────────────────────────────────────────────────────
# SBI — multi-sheet with Index sheet
# ─────────────────────────────────────────────────────────────
def import_sbi(cur, filepath: Path) -> tuple[int, int]:  # type: ignore
    """
    Index sheet: col[0]=code, col[1]=short, col[2]=name
    Fund sheets:
      Row 3 col[3] = fund name ("SCHEME NAME :" in col[2])
      Row 4 col[3] = date
      Row 6        = headers
      Row 8+       = data: name=col[2], isin=col[3], sector=col[4], weight=col[7]
    """
    try:
        wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
    except Exception as e:
        print(f"SBI read error: {e}")
        return 0, 0

    # Build index from "Index" sheet
    fund_index: dict[str, str] = {}
    if "Index" in wb.sheetnames:
        idx_ws = wb["Index"]
        for row in idx_ws.iter_rows(min_row=3, values_only=True):
            if row[1] and row[2]:
                fund_index[str(row[1]).strip()] = str(row[2]).strip()

    imported, skipped = 0, 0
    for sheet_name in wb.sheetnames:
        if sheet_name.lower() in ('index', 'cover'):
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(max_row=300, values_only=True))

        # Find fund name and date
        fund_name = fund_index.get(sheet_name)
        as_of_date = None
        header_row = -1

        for ri, row in enumerate(rows):
            cells = [str(c).strip() if c else '' for c in row]
            if any('scheme name' in c.lower() for c in cells):
                # Find the value after "SCHEME NAME :"
                for ci, c in enumerate(cells):
                    if 'scheme name' in c.lower() and ci + 1 < len(cells) and cells[ci + 1]:
                        fund_name = fund_name or cells[ci + 1]
                        break
            if any('portfolio statement' in c.lower() or 'as on' in c.lower() for c in cells):
                for ci, c in enumerate(cells):
                    if ('as on' in c.lower() or 'statement' in c.lower()) and ci + 1 < len(cells):
                        as_of_date = parse_date(cells[ci + 1]) or parse_date(c)
                        break
                if not as_of_date:
                    for c in cells:
                        if c:
                            as_of_date = parse_date(c)
                            if as_of_date:
                                break
            if any('name of the instrument' in c.lower() or 'instrument name' in c.lower() for c in cells):
                header_row = ri

        if header_row < 0:
            skipped += 1
            continue

        holdings = []
        for row in rows[header_row + 1:]:
            if len(row) < 8:
                continue
            # SBI: col[2]=name, col[3]=isin, col[4]=sector, col[7]=weight
            name   = clean_name(row[2])
            isin   = clean_name(row[3])
            sector = clean_name(row[4]) if len(row) > 4 else ''
            weight = to_float(row[7]) if len(row) > 7 else None

            if not is_valid_isin(isin) or weight is None or weight <= 0:
                continue
            holdings.append({'name': name, 'isin': isin, 'sector': sector, 'weight': weight})

        if not holdings:
            skipped += 1
            continue

        codes = match_scheme_codes(cur, fund_name or sheet_name)
        if not codes:
            print(f"  SBI '{fund_name}' — no DB match")
            skipped += 1
            continue
        print(f"  SBI '{fund_name}' → {len(codes)} variants, {len(holdings)} holdings, date={as_of_date}")
        for sc in codes:
            imported += upsert_holdings(cur, sc, holdings, as_of_date)

    wb.close()
    return imported, skipped


# ─────────────────────────────────────────────────────────────
# Quant — multi-sheet, fund name in row 2
# ─────────────────────────────────────────────────────────────
def import_quant(cur, filepath: Path) -> tuple[int, int]:  # type: ignore
    """
    Each sheet:
      Row 2 col[2] = fund name
      Row 4 col[2] = "MONTHLY PORTFOLIO STATEMENT AS ON ..."
      Row 8        = headers (col[0]=SR, col[1]=ISIN, col[2]=name, col[4]=sector, col[7]=weight)
      Row 9+       = data
    """
    try:
        wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
    except Exception as e:
        print(f"Quant read error: {e}")
        return 0, 0

    imported, skipped = 0, 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(max_row=300, values_only=True))

        fund_name = None
        as_of_date = None
        header_row = -1

        for ri, row in enumerate(rows[:10]):
            cells = [str(c).strip() if c else '' for c in row]
            if ri == 1 and len(cells) > 2 and cells[2] and 'quant' not in cells[2].lower():
                # Row 2 (index 1) has fund name in col 2
                if 'fund' in cells[2].lower() or 'scheme' in cells[2].lower():
                    fund_name = cells[2]
            if any('portfolio statement' in c.lower() or 'as on' in c.lower() for c in cells):
                for c in cells:
                    dt = parse_date(c)
                    if dt:
                        as_of_date = dt
                        break
            if 'name of the instrument' in ' '.join(cells).lower() or \
               ('isin' in cells and 'name' in ' '.join(cells).lower()):
                header_row = ri

        if header_row < 0:
            # Try to find headers at row 8
            for ri in range(5, min(12, len(rows))):
                cells = [str(c).strip() if c else '' for c in rows[ri]]
                if any('isin' in c.lower() for c in cells) and any('nav' in c.lower() or 'weight' in c.lower() or '%' in c for c in cells):
                    header_row = ri
                    break

        if header_row < 0:
            skipped += 1
            continue

        holdings = []
        for row in rows[header_row + 1:]:
            if len(row) < 8:
                continue
            # Quant: col[1]=ISIN, col[2]=name, col[4]=sector, col[7]=weight
            isin   = clean_name(row[1])
            name   = clean_name(row[2])
            sector = clean_name(row[4]) if len(row) > 4 else ''
            weight = to_float(row[7]) if len(row) > 7 else None

            if not is_valid_isin(isin) or weight is None or weight <= 0:
                continue
            holdings.append({'name': name, 'isin': isin, 'sector': sector, 'weight': weight})

        if not holdings:
            skipped += 1
            continue

        codes = match_scheme_codes(cur, fund_name or sheet_name)
        if not codes:
            print(f"  Quant '{fund_name or sheet_name}' — no DB match")
            skipped += 1
            continue
        print(f"  Quant '{fund_name}' → {len(codes)} variants, {len(holdings)} holdings, date={as_of_date}")
        for sc in codes:
            imported += upsert_holdings(cur, sc, holdings, as_of_date)

    wb.close()
    return imported, skipped


# ─────────────────────────────────────────────────────────────
# Kotak — multi-sheet, title in row 1
# ─────────────────────────────────────────────────────────────
def import_kotak(cur, filepath: Path) -> tuple[int, int]:  # type: ignore
    """
    Each sheet:
      Row 1 col[2] = "Portfolio of {Fund Name} as on ..."  OR  col[0] or col[2]
      Row 2        = headers (col[0]=name, col[3]=ISIN, col[4]=sector, col[8]=weight)
      Row 4+       = data (data starts at row 3 actually, col[2]=name, col[3]=ISIN, col[8]=%)
    """
    try:
        wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
    except Exception as e:
        print(f"Kotak read error: {e}")
        return 0, 0

    imported, skipped = 0, 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(max_row=300, values_only=True))
        if not rows:
            continue

        # Row 1: find fund name
        fund_name = None
        as_of_date = None
        row0 = [str(c).strip() if c else '' for c in rows[0]]
        for c in row0:
            m = re.search(r'Portfolio of (.+?)(?:\s+as on|\s*$)', c, re.IGNORECASE)
            if m:
                fund_name = m.group(1).strip()
            dt = parse_date(c)
            if dt:
                as_of_date = dt

        holdings = []
        for row in rows[2:]:
            if len(row) < 9:
                continue
            # Kotak col[2]=name, col[3]=ISIN, col[4]=sector, col[8]=weight
            name   = clean_name(row[2])
            isin   = clean_name(row[3])
            sector = clean_name(row[4]) if len(row) > 4 else ''
            weight = to_float(row[8]) if len(row) > 8 else None

            if not is_valid_isin(isin) or weight is None or weight <= 0:
                continue
            holdings.append({'name': name, 'isin': isin, 'sector': sector, 'weight': weight})

        if not holdings:
            skipped += 1
            continue

        codes = match_scheme_codes(cur, fund_name or sheet_name)
        if not codes:
            print(f"  Kotak '{fund_name or sheet_name}' — no DB match")
            skipped += 1
            continue
        print(f"  Kotak '{fund_name}' → {len(codes)} variants, {len(holdings)} holdings, date={as_of_date}")
        for sc in codes:
            imported += upsert_holdings(cur, sc, holdings, as_of_date)

    wb.close()
    return imported, skipped


# ─────────────────────────────────────────────────────────────
# Axis — multi-sheet with Index (same as import-amc-holdings.py)
# ─────────────────────────────────────────────────────────────
def import_axis(cur, filepath: Path) -> tuple[int, int]:  # type: ignore
    try:
        wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
    except Exception as e:
        print(f"Axis read error: {e}")
        return 0, 0

    # Build index
    fund_index: dict[str, str] = {}
    if "Index" in wb.sheetnames:
        idx_ws = wb["Index"]
        for row in idx_ws.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= 3 and row[1] and row[2]:
                fund_index[str(row[1]).strip()] = str(row[2]).strip()

    imported, skipped = 0, 0
    for sheet_name in wb.sheetnames:
        if sheet_name.lower() == 'index':
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(max_row=300, values_only=True))

        fund_name = fund_index.get(sheet_name)
        as_of_date = None
        header_row = -1

        for ri, row in enumerate(rows[:8]):
            cells = [str(c).strip() if c else '' for c in row]
            if any('as on' in c.lower() or 'statement date' in c.lower() for c in cells):
                for c in cells:
                    dt = parse_date(c)
                    if dt:
                        as_of_date = dt
                        break
            if any('name of instrument' in c.lower() or 'company name' in c.lower() for c in cells):
                header_row = ri

        if header_row < 0:
            skipped += 1
            continue

        holdings = []
        for row in rows[header_row + 1:]:
            if len(row) < 8:
                continue
            # Axis: col[2]=name, col[3]=ISIN, col[4]=sector, col[7]=weight
            name   = clean_name(row[2])
            isin   = clean_name(row[3])
            sector = clean_name(row[4]) if len(row) > 4 else ''
            weight = to_float(row[7]) if len(row) > 7 else None

            if not is_valid_isin(isin) or weight is None or weight <= 0:
                continue
            holdings.append({'name': name, 'isin': isin, 'sector': sector, 'weight': weight})

        if not holdings:
            skipped += 1
            continue

        codes = match_scheme_codes(cur, fund_name or sheet_name)
        if not codes:
            print(f"  Axis '{fund_name}' — no DB match")
            skipped += 1
            continue
        print(f"  Axis '{fund_name}' → {len(codes)} variants, {len(holdings)} holdings, date={as_of_date}")
        for sc in codes:
            imported += upsert_holdings(cur, sc, holdings, as_of_date)

    wb.close()
    return imported, skipped


# ─────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────

FILES = [
    ("axis",   BASE_DIR / "All-Schemes-Monthly-Portfolio---as-on-31st-January-2026.xlsx", import_axis),
    ("axis2",  BASE_DIR / "Axis.xlsx", import_axis),
    ("sbi",    BASE_DIR / "SBI_Feb2026.xlsx", import_sbi),
    ("quant",  BASE_DIR / "Quant_Feb2026.xlsx", import_quant),
    ("kotak",  BASE_DIR / "Kotak_Jan2026.xlsx", import_kotak),
    ("kotak2", BASE_DIR / "ConsolidatedSEBIPortfolioJanuary2026.xlsx", import_kotak),
]


def main():
    print(f"AMC Holdings Importer — {'DRY RUN' if DRY_RUN else 'LIVE'}")
    print(f"Base: {BASE_DIR}\n")

    conn = psycopg2.connect(DB_URL, connect_timeout=20)
    conn.autocommit = False
    cur = conn.cursor()

    # Enable trigram extension for fuzzy matching
    try:
        cur.execute("CREATE EXTENSION IF NOT EXISTS pg_trgm")
        conn.commit()
    except Exception:
        conn.rollback()

    total_imported = 0
    total_skipped  = 0

    # ICICI — per-fund files
    print("── ICICI ─────────────────────────────")
    imp, skp = import_icici(cur)
    total_imported += imp; total_skipped += skp
    print(f"   {imp} holdings imported, {skp} sheets skipped\n")

    # File-based AMCs
    for amc_id, filepath, handler in FILES:
        if not filepath.exists():
            print(f"── {amc_id.upper()} ── File not found: {filepath.name}, skipping")
            continue
        print(f"── {amc_id.upper()} ({filepath.name}) ──────────────────")
        imp, skp = handler(cur, filepath)
        total_imported += imp; total_skipped += skp
        print(f"   {imp} holdings imported, {skp} sheets skipped\n")

    if DRY_RUN:
        print("DRY RUN — rolling back")
        conn.rollback()
    else:
        conn.commit()
        print("Committed.")

    cur.close()
    conn.close()
    print(f"\n{'='*50}")
    print(f"Total: {total_imported} holdings imported, {total_skipped} sheets skipped")


if __name__ == "__main__":
    main()
