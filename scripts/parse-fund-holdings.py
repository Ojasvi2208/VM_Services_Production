"""
Configurable parser for SEBI monthly portfolio disclosure Excel files.
Reads AMC formats from amc-configs.json, parses top-N equity holdings per scheme,
and generates fund-holdings-parsed.json with diff report against previous run.

Usage: python3 scripts/parse-fund-holdings.py
"""

import openpyxl
import json
import os
import re
import sys
from collections import defaultdict

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_FILE = os.path.join(SCRIPT_DIR, "amc-configs.json")
OUTPUT_FILE = os.path.join(SCRIPT_DIR, "fund-holdings-parsed.json")
SQL_FILE = os.path.join(SCRIPT_DIR, "fund-holdings-insert.sql")


def load_config():
    with open(CONFIG_FILE, "r") as f:
        return json.load(f)


def load_previous():
    """Load previous parsed output for diff comparison."""
    if not os.path.exists(OUTPUT_FILE):
        return []
    with open(OUTPUT_FILE, "r") as f:
        return json.load(f)


def parse_index_sheet_format(filepath, amc_cfg, top_n=10):
    """Parse AMCs that use an index sheet mapping short_code -> scheme_name (e.g. SBI)."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    holdings = []

    idx_sheet = amc_cfg["index_sheet"]
    idx_ws = wb[idx_sheet]
    idx_rows = list(idx_ws.iter_rows(min_row=amc_cfg["index_start_row"], values_only=True))
    index_map = {}
    sc_col = amc_cfg["index_short_code_col"]
    sn_col = amc_cfg["index_scheme_name_col"]
    for row in idx_rows:
        if row[sc_col] and row[sn_col]:
            index_map[str(row[sc_col]).strip()] = str(row[sn_col]).strip()

    cols = amc_cfg["columns"]
    skip_names = set(s.upper() for s in amc_cfg.get("skip_stock_names", []))
    min_cols = amc_cfg["min_columns"]
    isin_prefix = amc_cfg["isin_prefix"]
    isin_len = amc_cfg["isin_length"]

    for sheet_name in wb.sheetnames:
        if sheet_name == idx_sheet:
            continue

        scheme_name = index_map.get(sheet_name, f"{amc_cfg['name']} Unknown ({sheet_name})")
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))

        equities = _extract_equities(all_rows, amc_cfg["data_start_row"], cols, min_cols, isin_prefix, isin_len, skip_names)
        holdings.extend(_rank_holdings(equities, amc_cfg, scheme_name, top_n))

    wb.close()
    return holdings


def parse_title_row_format(filepath, amc_cfg, top_n=10):
    """Parse AMCs where each sheet has a title row with fund name (e.g. Kotak)."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    holdings = []

    cols = amc_cfg["columns"]
    skip_names = set(s.upper() for s in amc_cfg.get("skip_stock_names", []))
    min_cols = amc_cfg["min_columns"]
    isin_prefix = amc_cfg["isin_prefix"]
    isin_len = amc_cfg["isin_length"]
    title_row_idx = amc_cfg["title_row"]
    title_col = amc_cfg["title_col"]
    title_regex = amc_cfg["title_regex"]

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))

        if len(all_rows) < 3:
            continue

        # Extract fund name from title row
        title_row = all_rows[title_row_idx]
        title_str = str(title_row[title_col]).strip() if len(title_row) > title_col and title_row[title_col] else ""
        match = re.match(title_regex, title_str)
        scheme_name = match.group(1).strip() if match else f"{amc_cfg['name']} ({sheet_name})"

        equities = _extract_equities(all_rows, amc_cfg["data_start_row"], cols, min_cols, isin_prefix, isin_len, skip_names)
        holdings.extend(_rank_holdings(equities, amc_cfg, scheme_name, top_n))

    wb.close()
    return holdings


def _extract_equities(all_rows, data_start_row, cols, min_cols, isin_prefix, isin_len, skip_names):
    """Extract equity holdings from rows using column config."""
    equities = []
    for row in all_rows[data_start_row:]:
        if len(row) < min_cols:
            continue

        isin = str(row[cols["isin"]]).strip() if row[cols["isin"]] else ""
        stock = str(row[cols["stock_name"]]).strip() if row[cols["stock_name"]] else ""
        weight = row[cols["weight_pct"]]
        sector = str(row[cols["sector"]]).strip() if cols.get("sector") is not None and row[cols["sector"]] else None

        if not isin or not isin.startswith(isin_prefix) or len(isin) < isin_len:
            continue
        if not stock or stock.upper() in skip_names or stock.strip() == "":
            continue

        try:
            w = float(weight)
        except (TypeError, ValueError):
            continue

        if w <= 0 or w > 100:
            continue

        equities.append({
            "stock_name": stock,
            "isin": isin[:isin_len],
            "sector": sector,
            "weight_pct": round(w, 2),
        })

    return equities


def _rank_holdings(equities, amc_cfg, scheme_name, top_n=10):
    """Sort by weight, take top N, return formatted holdings."""
    equities.sort(key=lambda x: x["weight_pct"], reverse=True)
    results = []
    for rank, eq in enumerate(equities[:top_n], 1):
        results.append({
            "amc": amc_cfg["name"],
            "scheme_name": scheme_name,
            "holding_name": eq["stock_name"],
            "holding_isin": eq["isin"],
            "sector": eq["sector"],
            "weight_pct": eq["weight_pct"],
            "rank": rank,
        })
    return results


PARSERS = {
    "index_sheet": parse_index_sheet_format,
    "title_row": parse_title_row_format,
}


def compute_diff(old_holdings, new_holdings):
    """Compare old vs new holdings and return a diff summary."""
    def make_key(h):
        return (h["amc"], h["scheme_name"], h["holding_name"], h["holding_isin"])

    old_set = {make_key(h): h for h in old_holdings}
    new_set = {make_key(h): h for h in new_holdings}

    old_keys = set(old_set.keys())
    new_keys = set(new_set.keys())

    added = new_keys - old_keys
    removed = old_keys - new_keys
    common = old_keys & new_keys

    weight_changes = []
    for k in common:
        old_w = old_set[k]["weight_pct"]
        new_w = new_set[k]["weight_pct"]
        if abs(old_w - new_w) >= 0.01:
            weight_changes.append({
                "amc": new_set[k]["amc"],
                "scheme": new_set[k]["scheme_name"],
                "holding": new_set[k]["holding_name"],
                "old_weight": old_w,
                "new_weight": new_w,
                "delta": round(new_w - old_w, 2),
            })

    old_schemes = set((h["amc"], h["scheme_name"]) for h in old_holdings)
    new_schemes = set((h["amc"], h["scheme_name"]) for h in new_holdings)

    return {
        "schemes_added": len(new_schemes - old_schemes),
        "schemes_removed": len(old_schemes - new_schemes),
        "holdings_added": len(added),
        "holdings_removed": len(removed),
        "weight_changes": len(weight_changes),
        "top_weight_changes": sorted(weight_changes, key=lambda x: abs(x["delta"]), reverse=True)[:20],
    }


def escape_sql(s):
    if s is None:
        return "NULL"
    return "'" + str(s).replace("'", "''") + "'"


def generate_sql(all_holdings, as_of_date):
    """Generate SQL file for direct DB ingestion."""
    with open(SQL_FILE, "w") as f:
        f.write(f"-- Auto-generated from SEBI portfolio disclosures ({as_of_date})\n")
        f.write("-- Parsed by parse-fund-holdings.py\n\n")

        f.write("""
CREATE TEMP TABLE IF NOT EXISTS staging_holdings (
  amc TEXT,
  scheme_name TEXT,
  holding_name TEXT,
  holding_isin VARCHAR(12),
  sector TEXT,
  weight_pct DECIMAL(5,2),
  rank INT
);

TRUNCATE staging_holdings;

""")

        batch_size = 200
        for i in range(0, len(all_holdings), batch_size):
            batch = all_holdings[i:i+batch_size]
            f.write("INSERT INTO staging_holdings (amc, scheme_name, holding_name, holding_isin, sector, weight_pct, rank) VALUES\n")
            values = []
            for h in batch:
                values.append(f"  ({escape_sql(h['amc'])}, {escape_sql(h['scheme_name'])}, {escape_sql(h['holding_name'])}, {escape_sql(h['holding_isin'])}, {escape_sql(h['sector'])}, {h['weight_pct']}, {h['rank']})")
            f.write(",\n".join(values))
            f.write(";\n\n")

        f.write(f"""
INSERT INTO fund_top_holdings (scheme_code, holding_name, holding_isin, weight_pct, sector, rank, as_of_date)
SELECT DISTINCT ON (f.scheme_code, sh.holding_isin)
  f.scheme_code,
  sh.holding_name,
  sh.holding_isin,
  sh.weight_pct,
  sh.sector,
  sh.rank,
  '{as_of_date}'::date
FROM staging_holdings sh
JOIN funds f ON (
  f.scheme_name ILIKE '%' || sh.scheme_name || '%'
  OR sh.scheme_name ILIKE '%' || f.scheme_name || '%'
  OR f.scheme_name ILIKE '%' || regexp_replace(sh.scheme_name, '^(SBI|Kotak)\\s+', '', 'i') || '%'
)
WHERE f.plan_type = 'Direct' AND f.option_type = 'Growth'
ON CONFLICT (scheme_code, holding_name, as_of_date) DO UPDATE SET
  holding_isin = EXCLUDED.holding_isin,
  weight_pct = EXCLUDED.weight_pct,
  sector = EXCLUDED.sector,
  rank = EXCLUDED.rank,
  fetched_at = NOW();

DROP TABLE IF EXISTS staging_holdings;

SELECT
  COUNT(DISTINCT scheme_code) as schemes_with_holdings,
  COUNT(*) as total_holdings
FROM fund_top_holdings
WHERE as_of_date = '{as_of_date}';
""")

    print(f"SQL written to {SQL_FILE}", flush=True)


def main():
    config = load_config()
    as_of_date = config["as_of_date"]
    previous = load_previous()

    all_holdings = []

    for amc_cfg in config["amcs"]:
        filepath = os.path.join(BASE, amc_cfg["file"])
        if not os.path.exists(filepath):
            print(f"SKIP {amc_cfg['name']}: file not found at {filepath}", flush=True)
            continue

        parser = PARSERS.get(amc_cfg["format"])
        if not parser:
            print(f"SKIP {amc_cfg['name']}: unknown format '{amc_cfg['format']}'", flush=True)
            continue

        print(f"Parsing {amc_cfg['name']}: {filepath}", flush=True)
        holdings = parser(filepath, amc_cfg, config.get("top_n", 10))
        all_holdings.extend(holdings)
        schemes = len(set(h["scheme_name"] for h in holdings))
        print(f"  {amc_cfg['name']}: {len(holdings)} holdings from {schemes} schemes", flush=True)

    total_schemes = len(set(h["scheme_name"] for h in all_holdings))
    print(f"\nTotal: {len(all_holdings)} holdings from {total_schemes} schemes", flush=True)

    # Diff against previous
    if previous:
        diff = compute_diff(previous, all_holdings)
        print(f"\n--- DIFF vs previous run ---")
        print(f"  Schemes added:    {diff['schemes_added']}")
        print(f"  Schemes removed:  {diff['schemes_removed']}")
        print(f"  Holdings added:   {diff['holdings_added']}")
        print(f"  Holdings removed: {diff['holdings_removed']}")
        print(f"  Weight changes:   {diff['weight_changes']}")
        if diff["top_weight_changes"]:
            print(f"  Top weight changes:")
            for wc in diff["top_weight_changes"][:10]:
                print(f"    {wc['amc']} | {wc['scheme'][:40]} | {wc['holding'][:25]} | {wc['old_weight']}% -> {wc['new_weight']}% ({wc['delta']:+.2f}%)")
    else:
        print("\nNo previous data — skipping diff.", flush=True)

    # Write output
    with open(OUTPUT_FILE, "w") as f:
        json.dump(all_holdings, f, indent=2)
    print(f"\nWritten to {OUTPUT_FILE}", flush=True)

    # Generate SQL
    generate_sql(all_holdings, as_of_date)


if __name__ == "__main__":
    main()
